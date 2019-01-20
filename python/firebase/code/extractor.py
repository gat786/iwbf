from openpyxl import load_workbook
import numpy as numpy


# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
def get_month_data(sheet_name=None):

    wb = load_workbook('python/firebase/code/abss.xlsx')
    sheet = wb['easy']
    if sheet_name!=None:
        sheet = wb[sheet_name]
    x = numpy.zeros(sheet.max_row)
    whole_month = []
    for column in sheet.columns:
        each_day = []
        for i in range(sheet.max_row):
            each_day.append(column[i].value)
        whole_month.append(each_day)
    print(whole_month)

def easyDay1Data():
    wb = load_workbook('day1easy.xlsx')
    sheet = wb['Sheet1']
    day1 = []
    for i in range(1,16):
        key = sheet['A'+str(i)].value
        reps = sheet['B'+str(i)].value
        duration = sheet['C'+str(i)].value
        sides = sheet['D'+str(i)].value
        day1.append({'key':key,'reps':reps,'duration':duration,'sides':sides})
    return day1
easyDay1Data()